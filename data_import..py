import sqlite3
import pandas as pd
import os
from openpyxl.formatting.rule import ColorScaleRule

os.makedirs("exports", exist_ok=True)

conn = sqlite3.connect("soccer.db")

# Matches per league
query1 = """
SELECT L.name AS league_name, COUNT(M.id) AS matches_count
FROM Match M
JOIN League L ON M.league_id = L.id
GROUP BY L.name;
"""
df1 = pd.read_sql(query1, conn)

# Average goals per season
query2 = """
SELECT substr(M.date, 1, 4) AS year,
       AVG(M.home_team_goal + M.away_team_goal) AS avg_goals
FROM Match M
WHERE substr(M.date, 1, 4) IS NOT NULL
GROUP BY year
ORDER BY year;
"""
df2 = pd.read_sql(query2, conn)

# Top 15 teams by matches played
query3 = """
SELECT T.team_long_name AS team_name, COUNT(M.id) AS matches_played
FROM Team T
JOIN Match M ON T.team_api_id = M.home_team_api_id OR T.team_api_id = M.away_team_api_id
GROUP BY T.team_long_name
ORDER BY matches_played DESC
LIMIT 15;
"""
df3 = pd.read_sql(query3, conn)
conn.close()

dfs = {
    "Leagues": df1,
    "GoalsPerSeason": df2,
    "TopTeams": df3
}

def export_to_excel(dataframes_dict, filename):
    filepath = os.path.join("exports", filename)

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        total_rows = 0
        for sheet, df in dataframes_dict.items():
            df.to_excel(writer, sheet_name=sheet, index=False)
            total_rows += len(df)

        workbook = writer.book
        for sheet in workbook.sheetnames:
            ws = workbook[sheet]
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            for col in ws.iter_cols(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                if all(isinstance(cell.value, (int, float, type(None))) for cell in col):
                    col_letter = col[0].column_letter
                    rule = ColorScaleRule(
                        start_type="min", start_color="FFAA0000",
                        mid_type="percentile", mid_value=50, mid_color="FFFFFF00",
                        end_type="max", end_color="FF00AA00"
                    )
                    ws.conditional_formatting.add(f"{col_letter}2:{col_letter}{ws.max_row}", rule)

    print(f"Created file {filename}, {len(dataframes_dict)} sheets, {total_rows} rows")

export_to_excel(dfs, "soccer_report.xlsx")
